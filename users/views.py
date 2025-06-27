from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth import logout
from .forms import CustomAuthenticationForm 
from compliance.models import Employee
from compensation.models import EmployeeCTC
import openpyxl
from django.http import HttpResponse
from users.models import Profile
from applicant.models import Applicant
from django.utils.dateparse import parse_date
from django.db.models import Q
from io import BytesIO
import openpyxl
from .models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference

def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            login(request, form.user_cache)
            try:
                if request.user.applicant:  # Trying to access the related applicant
                    return redirect('applicant:dashboard')
            except request.user._meta.model.applicant.RelatedObjectDoesNotExist:
                return redirect('core:main_page')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'users/login.html', {'form': form})

def logout_view(request):
    logout(request)

    return redirect('users:login_view')

########################### Modifying code Personal Information for columns wise

def export_profile_to_excel(profile, family_members, educational_qualifications, employment_records, language_proficiencies, references, compliance_info, compensation_info):
    import openpyxl
    from openpyxl.styles import Alignment
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Profile Details"

    # Define headers for the personal details section
    personal_headers = [
        "Employee Code", "First Name", "Middle Name", "Last Name",
        "Designation", "Date of Joining", "End Date", "Email",
        "Phone Number", "Gender", "Date of Birth", "Present Address",
        "Permanent Address", "Personal Email", "PAN Number", "Marital Status"
    ]
    personal_details = [
        profile.employee_code, profile.first_name, profile.middle_name,
        profile.last_name, profile.designation, profile.date_of_joining,
        profile.end_date, profile.email, profile.phone_number,
        profile.get_gender_display(), profile.date_of_birth,
        profile.present_address, profile.permanent_address,
        profile.personal_email, profile.pan_number, profile.marital_status
    ]
    # Write headers and personal details in column-wise format
    for col, header in enumerate(personal_headers, start=1):
        sheet.cell(row=1, column=col, value=header)  # Header row
        sheet.cell(row=2, column=col, value=personal_details[col - 1])  # Data row

    # Adjust alignment for readability
    for col in range(1, len(personal_headers) + 1):
        sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        sheet.cell(row=2, column=col).alignment = Alignment(horizontal='center')

    # Add additional sections for family members, education, etc.
    row_offset = 4

    # Family Members Section
    sheet.cell(row=row_offset, column=1, value="Family Members")
    row_offset += 1
    sheet.append(["Relation", "Name", "Date of Birth", "Sex", "Age"])
    for member in family_members:
        sheet.append([member.relation, member.name, member.date_of_birth, member.sex, member.age])
    row_offset += len(family_members) + 2

    # Educational Qualifications
    sheet.cell(row=row_offset, column=1, value="Educational Qualifications")
    row_offset += 1
    sheet.append(["Examination Passed", "Year of Passing", "School/College", "Division"])
    for qualification in educational_qualifications:
        sheet.append([
            qualification.examination_passed,
            qualification.year_of_passing,
            qualification.school_or_college,
            qualification.division,
        ])
    row_offset += len(educational_qualifications) + 2

    # Employment Records
    sheet.cell(row=row_offset, column=1, value="Employment Records")
    row_offset += 1
    sheet.append(["Organization", "Designation", "Joining Date", "Leaving Date"])
    for record in employment_records:
        sheet.append([
            record.organization, record.designation,
            record.joining_date, record.leaving_date
        ])
    row_offset += len(employment_records) + 2

    # Language Proficiencies
    sheet.cell(row=row_offset, column=1, value="Languages Known")
    row_offset += 1
    sheet.append(["Language", "Speak", "Read", "Write"])
    for language in language_proficiencies:
        sheet.append([
            language.language,
            language.get_speak_display(),
            language.get_read_display(),
            language.get_write_display()
        ])
    row_offset += len(language_proficiencies) + 2

    # References
    sheet.cell(row=row_offset, column=1, value="References")
    row_offset += 1
    sheet.append(["Name", "Occupation", "Phone Number", "Email"])
    for reference in references:
        sheet.append([
            reference.name, reference.occupation,
            reference.phone_number, reference.email
        ])
    row_offset += len(references) + 2

    # Compliance Information
    sheet.cell(row=row_offset, column=1, value="Compliance Information")
    row_offset += 1
    if compliance_info:
        sheet.append(["UAN Number", compliance_info.uan_number])
        sheet.append(["PAN Number", compliance_info.pan_number])
        sheet.append(["ESIC Number", compliance_info.esic_number])
    else:
        sheet.append(["No compliance data available"])
    row_offset += 3

    # Compensation Information
    sheet.cell(row=row_offset, column=1, value="Compensation Information")
    row_offset += 1
    if compensation_info:
        sheet.append(["Component", "Amount"])
        salary = compensation_info.salary
        contribution = compensation_info.contribution
        deduction = compensation_info.deduction
        sheet.append(["Basic Salary", salary.basic])
        sheet.append(["HRA", salary.hra])
        sheet.append(["Special Allowance", salary.special_allowance])
        sheet.append(["Gross Salary", salary.gross_salary()])
        sheet.append(["Total Contributions", contribution.emp_contribution()])
        sheet.append(["Total Deductions", deduction.emp_deduction()])
        sheet.append(["Net Salary (CTC)", compensation_info.calculate_ctc()])
    else:
        sheet.append(["No compensation data available"])

    return workbook

################################ update views for compliance data
def profile_detail_view(request, profile_id):
    # Get the specific profile by its ID
    profile = get_object_or_404(Profile, id=profile_id)
    all_profiles = Profile.objects.all()

    # Access related FamilyMember, EducationalQualification, etc.
    family_members = profile.family_members.all()
    educational_qualifications = profile.educational_qualifications.all()
    employment_records = profile.employment_records.all()
    language_proficiencies = profile.languages.all()
    references = profile.references.all()

    # Fetch the compliance data through the Applicant linked to this Profile
    applicant = getattr(profile.user, 'applicant', None)  # Access Applicant via the CustomUser relationship
    compliance_info = Employee.objects.filter(applicant=applicant).first() if applicant else None
    compensation_info = EmployeeCTC.objects.filter(applicant=applicant).first()
    # compensation_info = EmployeeCTC.objects.filter(profile=profile).first()

    # Export to Excel if requested
    if request.GET.get('export') == 'excel':
        workbook = export_profile_to_excel(
            profile,
            family_members,
            educational_qualifications,
            employment_records,
            language_proficiencies,
            references,
            compliance_info,
            compensation_info  # Add these two arguments
        )
    # Prepare the response
        response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Profile_{profile.employee_code}.xlsx'
        workbook.save(response)
        return response
    # Pass the data to the template if not exporting
    context = {
        'profile': profile,
        'all_profiles': all_profiles,
        'family_members': family_members,
        'educational_qualifications': educational_qualifications,
        'employment_records': employment_records,
        'language_proficiencies': language_proficiencies,
        'references': references,
        'compliance_info': compliance_info,
        'compensation_info': compensation_info,
    }

    return render(request, 'profile_detail.html', context)

def hiring_report(request):
    from_date = request.GET.get('from_date')
    date_of_joining = request.GET.get('date_of_joining')
    
    profiles = Profile.objects.all()
    
    # Apply the filter based on 'date_of_joining' range
    if from_date and date_of_joining:
        profiles = profiles.filter(
            date_of_joining__range=(parse_date(from_date), parse_date(date_of_joining))
        )
    elif from_date:
        profiles = profiles.filter(date_of_joining__gte=parse_date(from_date))
    elif date_of_joining:
        profiles = profiles.filter(date_of_joining__lte=parse_date(date_of_joining))
    
    return render(request, 'users/hiring_report.html', {'profiles': profiles, 'from_date': from_date, 'date_of_joining': date_of_joining})


def attrition_report(request):
    from_date = request.GET.get('from_date')
    end_date = request.GET.get('end_date')
    
    profiles = Profile.objects.filter(end_date__isnull=False)  # Only include profiles with an end_date

    # Apply the filter based only on the 'end_date' range
    if from_date and end_date:
        profiles = profiles.filter(
            end_date__range=(parse_date(from_date), parse_date(end_date))
        )
    elif from_date:
        profiles = profiles.filter(end_date__gte=parse_date(from_date))
    elif end_date:
        profiles = profiles.filter(end_date__lte=parse_date(end_date))
    
    return render(request, 'users/attrition_report.html', {'profiles': profiles, 'from_date': from_date, 'end_date': end_date})


########################### export all applicant details
########################## Add compensation and complianc  #################

from django.http import HttpResponse
from openpyxl import Workbook
from applicant.models import Applicant
from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference


# def export_applicants_to_excel(request):
#     # Create a new Excel workbook and active sheet
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"

#     # Define the headers for the Excel sheet
#     headers = [
#         "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date", 
#         "Gender", "Date of Birth", "Age", "Email", "Phone Number", "Personal Email", 
#         "Father Name", "Mother Name", "Blood Group", "Present Address", "Permanent Address", 
#         "Marital Status", "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number", 
#         "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",
#         "Family Members", "Educational Qualifications", "Employment Records", "Language Proficiencies", 
#         "References", "UAN Number", "PAN Number (Compliance)", "ESIC Number", "Gross Salary", 
#         "Total Contribution", "Total Deductions", "Net CTC"
#     ]

#     # Write headers to the first row of the sheet
#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=1, column=col_idx, value=header)

#     # Fetch applicants and write their data row by row
#     applicants = Applicant.objects.select_related('profile', 'user')
#     for row_idx, applicant in enumerate(applicants, start=2):
#         profile = applicant.profile

#         # Write profile data
#         sheet.cell(row=row_idx, column=1, value=profile.fullname)
#         sheet.cell(row=row_idx, column=2, value=profile.employee_code)
#         sheet.cell(row=row_idx, column=3, value=profile.designation)
#         sheet.cell(row=row_idx, column=4, value=profile.date_of_joining)
#         sheet.cell(row=row_idx, column=5, value=profile.end_date)
#         sheet.cell(row=row_idx, column=6, value=profile.get_gender_display())
#         sheet.cell(row=row_idx, column=7, value=profile.date_of_birth)
#         sheet.cell(row=row_idx, column=8, value=profile.age)
#         sheet.cell(row=row_idx, column=9, value=profile.email)
#         sheet.cell(row=row_idx, column=10, value=profile.phone_number)
#         sheet.cell(row=row_idx, column=11, value=profile.personal_email)
#         sheet.cell(row=row_idx, column=12, value=profile.father_name)
#         sheet.cell(row=row_idx, column=13, value=profile.mother_name)
#         sheet.cell(row=row_idx, column=14, value=profile.blood_group)
#         sheet.cell(row=row_idx, column=15, value=profile.present_address)
#         sheet.cell(row=row_idx, column=16, value=profile.permanent_address)
#         sheet.cell(row=row_idx, column=17, value=profile.marital_status)
#         sheet.cell(row=row_idx, column=18, value=profile.pan_number)
#         sheet.cell(row=row_idx, column=19, value=profile.bank_name)
#         sheet.cell(row=row_idx, column=20, value=profile.ifsc_code)
#         sheet.cell(row=row_idx, column=21, value=profile.bank_account_number)
#         sheet.cell(row=row_idx, column=22, value=profile.emergency_contact_name)
#         sheet.cell(row=row_idx, column=23, value=profile.emergency_contact_relation)
#         sheet.cell(row=row_idx, column=24, value=profile.emergency_contact_number)

#         # Write related data as concatenated strings
#         # Family Members
#         family_members = ", ".join([f"{fm.name} ({fm.relation})" for fm in profile.family_members.all()])
#         sheet.cell(row=row_idx, column=25, value=family_members)

#         # Educational Qualifications
#         educational_qualifications = ", ".join([
#             f"{eq.examination_passed} ({eq.year_of_passing})" for eq in profile.educational_qualifications.all()
#         ])
#         sheet.cell(row=row_idx, column=26, value=educational_qualifications)

#         # Employment Records
#         employment_records = ", ".join([
#             f"{er.organization} ({er.designation})" for er in profile.employment_records.all()
#         ])
#         sheet.cell(row=row_idx, column=27, value=employment_records)

#         # Language Proficiencies
#         languages = ", ".join([
#             f"{lp.language} (Speak: {lp.speak}, Read: {lp.read}, Write: {lp.write})" 
#             for lp in profile.languages.all()
#         ])
#         sheet.cell(row=row_idx, column=28, value=languages)

#         # References
#         references = ", ".join([
#             f"{ref.name} ({ref.occupation}, {ref.phone_number})" for ref in profile.references.all()
#         ])
#         sheet.cell(row=row_idx, column=29, value=references)

#         # Add Compliance details
#         try:
#             compliance = applicant.employee
#             sheet.cell(row=row_idx, column=30, value=compliance.uan_number)
#             sheet.cell(row=row_idx, column=31, value=compliance.pan_number)
#             sheet.cell(row=row_idx, column=32, value=compliance.esic_number)
#         except AttributeError:
#             # If no compliance record exists
#             sheet.cell(row=row_idx, column=30, value="N/A")
#             sheet.cell(row=row_idx, column=31, value="N/A")
#             sheet.cell(row=row_idx, column=32, value="N/A")

#         # Add Compensation details
#         try:
#             compensation = applicant.employeectc
#             sheet.cell(row=row_idx, column=33, value=compensation.salary.gross_salary())
#             sheet.cell(row=row_idx, column=34, value=compensation.contribution.emp_contribution())
#             sheet.cell(row=row_idx, column=35, value=compensation.deduction.emp_deduction())
#             sheet.cell(row=row_idx, column=36, value=compensation.calculate_ctc())
#         except AttributeError:
#             # If no compensation record exists
#             sheet.cell(row=row_idx, column=33, value="N/A")
#             sheet.cell(row=row_idx, column=34, value="N/A")
#             sheet.cell(row=row_idx, column=35, value="N/A")
#             sheet.cell(row=row_idx, column=36, value="N/A")

#     # Set the HTTP response with the Excel file
#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response


   


def export_applicants_to_excel(request):
    # Create a new Excel workbook and active sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Applicants Data"

    # Define the headers for the Excel sheet
    headers = [
        "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date",
        "Gender", "Date of Birth", "Age", "Email", "Phone Number", "Personal Email",
        "Father Name", "Mother Name", "Blood Group", "Present Address", "Permanent Address",
        "Marital Status", "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number",
        "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",
        "Family Members", "Educational Qualifications", "Employment Records", "Language Proficiencies",
        "References", "UAN Number", "PAN Number (Compliance)", "ESIC Number", "Gross Salary",
        "Total Contribution", "Total Deductions", "Net CTC"
    ]

    # Write headers to the first row of the sheet
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Fetch applicants and related data
    applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
        'profile__family_members',
        'profile__educational_qualifications',
        'profile__employment_records',
        'profile__languages',
        'profile__references',
    )

    for row_idx, applicant in enumerate(applicants, start=2):
        profile = applicant.profile

        # Write profile data
        sheet.cell(row=row_idx, column=1, value=profile.fullname)
        sheet.cell(row=row_idx, column=2, value=profile.employee_code)
        sheet.cell(row=row_idx, column=3, value=profile.designation)
        sheet.cell(row=row_idx, column=4, value=profile.date_of_joining)
        sheet.cell(row=row_idx, column=5, value=profile.end_date)
        sheet.cell(row=row_idx, column=6, value=profile.get_gender_display())
        sheet.cell(row=row_idx, column=7, value=profile.date_of_birth)
        sheet.cell(row=row_idx, column=8, value=profile.age)
        sheet.cell(row=row_idx, column=9, value=profile.email)
        sheet.cell(row=row_idx, column=10, value=profile.phone_number)
        sheet.cell(row=row_idx, column=11, value=profile.personal_email)
        sheet.cell(row=row_idx, column=12, value=profile.father_name)
        sheet.cell(row=row_idx, column=13, value=profile.mother_name)
        sheet.cell(row=row_idx, column=14, value=profile.blood_group)
        sheet.cell(row=row_idx, column=15, value=profile.present_address)
        sheet.cell(row=row_idx, column=16, value=profile.permanent_address)
        sheet.cell(row=row_idx, column=17, value=profile.marital_status)
        sheet.cell(row=row_idx, column=18, value=profile.pan_number)
        sheet.cell(row=row_idx, column=19, value=profile.bank_name)
        sheet.cell(row=row_idx, column=20, value=profile.ifsc_code)
        sheet.cell(row=row_idx, column=21, value=profile.bank_account_number)
        sheet.cell(row=row_idx, column=22, value=profile.emergency_contact_name)
        sheet.cell(row=row_idx, column=23, value=profile.emergency_contact_relation)
        sheet.cell(row=row_idx, column=24, value=profile.emergency_contact_number)

        # Write related data
        family_members = ", ".join([f"{fm.name} ({fm.relation})" for fm in profile.family_members.all()])
        sheet.cell(row=row_idx, column=25, value=family_members)

        educational_qualifications = ", ".join([
            f"{eq.examination_passed} ({eq.year_of_passing})" for eq in profile.educational_qualifications.all()
        ])
        sheet.cell(row=row_idx, column=26, value=educational_qualifications)

        employment_records = ", ".join([
            f"{er.organization} ({er.designation})" for er in profile.employment_records.all()
        ])
        sheet.cell(row=row_idx, column=27, value=employment_records)

        languages = ", ".join([
            f"{lp.language} (Speak: {lp.speak}, Read: {lp.read}, Write: {lp.write})"
            for lp in profile.languages.all()
        ])
        sheet.cell(row=row_idx, column=28, value=languages)

        references = ", ".join([
            f"{ref.name} ({ref.occupation}, {ref.phone_number})" for ref in profile.references.all()
        ])
        sheet.cell(row=row_idx, column=29, value=references)

        # Compliance details
        try:
            compliance = applicant.employee
            sheet.cell(row=row_idx, column=30, value=compliance.uan_number)
            sheet.cell(row=row_idx, column=31, value=compliance.pan_number)
            sheet.cell(row=row_idx, column=32, value=compliance.esic_number)
        except AttributeError:
            sheet.cell(row=row_idx, column=30, value="N/A")
            sheet.cell(row=row_idx, column=31, value="N/A")
            sheet.cell(row=row_idx, column=32, value="N/A")

        # Compensation details
        try:
            compensation = EmployeeCTC.objects.get(applicant=applicant)
            sheet.cell(row=row_idx, column=33, value=compensation.salary.gross_salary())
            sheet.cell(row=row_idx, column=34, value=compensation.contribution.emp_contribution())
            sheet.cell(row=row_idx, column=35, value=compensation.deduction.emp_deduction())
            sheet.cell(row=row_idx, column=36, value=compensation.calculate_ctc())
        except EmployeeCTC.DoesNotExist:
            sheet.cell(row=row_idx, column=33, value="N/A")
            sheet.cell(row=row_idx, column=34, value="N/A")
            sheet.cell(row=row_idx, column=35, value="N/A")
            sheet.cell(row=row_idx, column=36, value="N/A")

    # Set the HTTP response with the Excel file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
    workbook.save(response)
    return response
