from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib.auth import logout
from .forms import CustomAuthenticationForm 
from django.contrib.auth import login
from django.shortcuts import redirect

from compliance.models import Employee


from compensation.models import EmployeeCTC


def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            login(request, form.user_cache)
            try:
                if request.user.applicant:  # Trying to access the related applicant
                    return redirect('applicant:dashboard')
            except request.user._meta.model.applicant.RelatedObjectDoesNotExist:
                return redirect('core:main_page')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'users/login.html', {'form': form})



def logout_view(request):
    logout(request)

    return redirect('users:login_view')





# from django.shortcuts import render, get_object_or_404
# from users.models import Profile

# def profile_detail_view(request, profile_id):
#     # Get the specific profile by its ID
#     profile = get_object_or_404(Profile, id=profile_id)

#     # Access related FamilyMember, EducationalQualification, etc.
#     family_members = profile.family_members.all()
#     educational_qualifications = profile.educational_qualifications.all()
#     employment_records = profile.employment_records.all()
#     language_proficiencies = profile.languages.all()
#     references = profile.references.all()

#     # Pass the data to the template
#     context = {
#         'profile': profile,
#         'family_members': family_members,
#         'educational_qualifications': educational_qualifications,
#         'employment_records': employment_records,
#         'language_proficiencies': language_proficiencies,
#         'references': references,
#     }

#     return render(request, 'profile_detail.html', context)



#################### Different pages

# import openpyxl
# from django.http import HttpResponse
# from django.shortcuts import render, get_object_or_404
# from users.models import Profile

# def profile_detail_view(request, profile_id):
#     # Get the specific profile by its ID
#     profile = get_object_or_404(Profile, id=profile_id)

#     # Access related FamilyMember, EducationalQualification, etc.
#     family_members = profile.family_members.all()
#     educational_qualifications = profile.educational_qualifications.all()
#     employment_records = profile.employment_records.all()
#     language_proficiencies = profile.languages.all()
#     references = profile.references.all()

#     # If the request is for exporting the data to Excel
#     if request.GET.get('export') == 'excel':
#         return export_to_excel(profile, family_members, educational_qualifications, employment_records, language_proficiencies, references)

#     # Pass the data to the template
#     context = {
#         'profile': profile,
#         'family_members': family_members,
#         'educational_qualifications': educational_qualifications,
#         'employment_records': employment_records,
#         'language_proficiencies': language_proficiencies,
#         'references': references,
#     }

#     return render(request, 'profile_detail.html', context)

# def export_to_excel(profile, family_members, educational_qualifications, employment_records, language_proficiencies, references):
#     # Create an in-memory Excel file
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Profile Details"

#     # Write the profile details to the Excel sheet
#     ws.append(["Field", "Details"])
#     ws.append(["Employee Code", profile.Employee_Code])
#     ws.append(["First Name", profile.first_name])
#     ws.append(["Middle Name", profile.middle_name])
#     ws.append(["Last Name", profile.last_name])
#     ws.append(["Designation", profile.designation])
#     ws.append(["Date of Joining", profile.date_of_joining])
#     ws.append(["Email", profile.email])
#     ws.append(["Phone Number", profile.phone_number])
#     ws.append(["Gender", profile.get_gender_display()])
#     ws.append(["Date of Birth", profile.date_of_birth])
#     ws.append(["Present Address", profile.present_address])
#     ws.append(["Permanent Address", profile.permanent_address])
#     ws.append(["Personal Email", profile.personal_email])
#     ws.append(["PAN Number", profile.pan_number])
#     ws.append(["Marital Status", profile.marital_status])

#     # Add a new sheet for Family Members
#     family_ws = wb.create_sheet(title="Family Members")
#     family_ws.append(["Relation", "Name", "Date of Birth", "Sex", "Age"])
#     for member in family_members:
#         family_ws.append([member.relation, member.name, member.date_of_birth, member.sex, member.age])

#     # Add a new sheet for Educational Qualifications
#     education_ws = wb.create_sheet(title="Educational Qualifications")
#     education_ws.append(["Examination Passed", "Year of Passing", "School/College", "Division", "Document"])
#     for qualification in educational_qualifications:
#         education_ws.append([qualification.examination_passed, qualification.year_of_passing, qualification.school_or_college, qualification.division, qualification.document.url if qualification.document else "No Document"])

#     # Add a new sheet for Employment Records
#     employment_ws = wb.create_sheet(title="Employment Records")
#     employment_ws.append(["Organization", "Designation", "Joining Date", "Leaving Date", "Document"])
#     for record in employment_records:
#         employment_ws.append([record.organization, record.designation, record.joining_date, record.leaving_date, record.document.url if record.document else "No Document"])

#     # Add a new sheet for Language Proficiencies
#     language_ws = wb.create_sheet(title="Languages Known")
#     language_ws.append(["Language", "Speak", "Read", "Write"])
#     for language in language_proficiencies:
#         language_ws.append([language.language, language.get_speak_display(), language.get_read_display(), language.get_write_display()])

#     # Add a new sheet for References
#     reference_ws = wb.create_sheet(title="References")
#     reference_ws.append(["Name", "Occupation", "Phone Number", "Email"])
#     for reference in references:
#         reference_ws.append([reference.name, reference.occupation, reference.phone_number, reference.email])

#     # Prepare the response as an Excel file
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=profile_details.xlsx'
#     wb.save(response)

#     return response



######################## Single page



# import openpyxl
# from django.http import HttpResponse
# from django.shortcuts import render, get_object_or_404
# from users.models import Profile

# def export_profile_to_excel(profile, family_members, educational_qualifications, employment_records, language_proficiencies, references):
#     # Create an Excel workbook and worksheet

#     all_profiles = Profile.objects.all()

#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Profile Details"

#     # Define headers for the personal details section
#     sheet.append(["Field", "Details"])
#     sheet.append(["Employee Code", profile.Employee_Code])
#     sheet.append(["First Name", profile.first_name])
#     sheet.append(["Middle Name", profile.middle_name])
#     sheet.append(["Last Name", profile.last_name])
#     sheet.append(["Designation", profile.designation])
#     sheet.append(["Date of Joining", profile.date_of_joining])
#     sheet.append(["Email", profile.email])
#     sheet.append(["Phone Number", profile.phone_number])
#     sheet.append(["Gender", profile.get_gender_display()])
#     sheet.append(["Date of Birth", profile.date_of_birth])
#     sheet.append(["Present Address", profile.present_address])
#     sheet.append(["Permanent Address", profile.permanent_address])
#     sheet.append(["Personal Email", profile.personal_email])
#     sheet.append(["PAN Number", profile.pan_number])
#     sheet.append(["Marital Status", profile.marital_status])

#     # Adding a blank row
#     sheet.append([])

#     # Family members section
#     sheet.append(["Family Members"])
#     sheet.append(["Relation", "Name", "Date of Birth", "Sex", "Age"])
#     for member in family_members:
#         sheet.append([member.relation, member.name, member.date_of_birth, member.sex, member.age])

#     sheet.append([])

#     # Educational qualifications section
#     sheet.append(["Educational Qualifications"])
#     sheet.append(["Examination Passed", "Year of Passing", "School/College", "Division"])
#     for qualification in educational_qualifications:
#         sheet.append([qualification.examination_passed, qualification.year_of_passing, qualification.school_or_college, qualification.division])

#     sheet.append([])

#     # Employment records section
#     sheet.append(["Employment Records"])
#     sheet.append(["Organization", "Designation", "Joining Date", "Leaving Date"])
#     for record in employment_records:
#         sheet.append([record.organization, record.designation, record.joining_date, record.leaving_date])

#     sheet.append([])

#     # Language proficiencies section
#     sheet.append(["Language Proficiencies"])
#     sheet.append(["Language", "Speak", "Read", "Write"])
#     for language in language_proficiencies:
#         sheet.append([language.language, language.get_speak_display(), language.get_read_display(), language.get_write_display()])

#     sheet.append([])

#     # References section
#     sheet.append(["References"])
#     sheet.append(["Name", "Occupation", "Phone Number", "Email"])
#     for reference in references:
#         sheet.append([reference.name, reference.occupation, reference.phone_number, reference.email])

#     return workbook



###################### changes 1



import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from users.models import Profile
# from employee.models import EmployeeCTC, Employee  # Assuming EmployeeCTC and Employee are in employee.models

def export_profile_to_excel(profile, family_members, educational_qualifications, employment_records, language_proficiencies, references, compliance_info, compensation_info):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Profile Details"

    # Define headers for the personal details section
    sheet.append(["Field", "Details"])
    sheet.append(["Employee Code", profile.Employee_Code])
    sheet.append(["First Name", profile.first_name])
    sheet.append(["Middle Name", profile.middle_name])
    sheet.append(["Last Name", profile.last_name])
    sheet.append(["Designation", profile.designation])
    sheet.append(["Date of Joining", profile.date_of_joining])
    sheet.append(["Email", profile.email])
    sheet.append(["Phone Number", profile.phone_number])
    sheet.append(["Gender", profile.get_gender_display()])
    sheet.append(["Date of Birth", profile.date_of_birth])
    sheet.append(["Present Address", profile.present_address])
    sheet.append(["Permanent Address", profile.permanent_address])
    sheet.append(["Personal Email", profile.personal_email])
    sheet.append(["PAN Number", profile.pan_number])
    sheet.append(["Marital Status", profile.marital_status])

    # Adding a blank row
    sheet.append([])

    # Compliance section
    sheet.append(["Compliance Information"])
    if compliance_info:
        sheet.append(["Compliance Field", "Compliance Value"])  # Add relevant fields here
        # sheet.append(["Compliance Status", compliance_info.status])  # Example field
        # sheet.append(["UAN number", Employee.uan_number])
        # sheet.append(["PAN Number", Employee.pan_number])
        # sheet.append(["ESIC Number", Employee.esic_number])
        sheet.append(["Compliance Status", getattr(compliance_info, 'status', 'N/A')])
        sheet.append(["UAN Number", getattr(compliance_info, 'uan_number', 'N/A')])
        sheet.append(["PAN Number", getattr(compliance_info, 'pan_number', 'N/A')])
        sheet.append(["ESIC Number", getattr(compliance_info, 'esic_number', 'N/A')])
    else:
        sheet.append(["No compliance data available"])

    sheet.append([])

    # Compensation section
    sheet.append(["Compensation Information"])
    # if compensation_info:
    #     sheet.append(["Component", "Amount"])
    #     sheet.append(["Basic Salary", compensation_info.basic_salary])  # Example fields
    #     sheet.append(["Allowances", compensation_info.allowances])
    #     sheet.append(["Deductions", compensation_info.deductions])
    #     sheet.append(["Net Salary", compensation_info.net_salary])
                


    if compensation_info:
        sheet.append(["Component", "Amount"])
        salary = compensation_info.salary
        contribution = compensation_info.contribution
        deduction = compensation_info.deduction
        sheet.append(["Basic Salary", salary.basic])
        sheet.append(["HRA", salary.hra])
        sheet.append(["Special Allowance", salary.special_allowance])
        sheet.append(["Gross Salary", salary.gross_salary()])
        sheet.append(["Total Contributions", contribution.emp_contribution()])
        sheet.append(["Total Deductions", deduction.emp_deduction()])
        sheet.append(["Net Salary (CTC)", compensation_info.calculate_ctc()])


    else:
        sheet.append(["No compensation data available"])

    # Adding additional sections such as family members, educational qualifications, etc.
    sheet.append([])
    sheet.append(["Family Members"])
    sheet.append(["Relation", "Name", "Date of Birth", "Sex", "Age"])
    for member in family_members:
        sheet.append([member.relation, member.name, member.date_of_birth, member.sex, member.age])

    sheet.append([])
    sheet.append(["Educational Qualifications"])
    sheet.append(["Examination Passed", "Year of Passing", "School/College", "Division"])
    for qualification in educational_qualifications:
        sheet.append([qualification.examination_passed, qualification.year_of_passing, qualification.school_or_college, qualification.division])

    sheet.append([])
    sheet.append(["Employment Records"])
    sheet.append(["Organization", "Designation", "Joining Date", "Leaving Date"])
    for record in employment_records:
        sheet.append([record.organization, record.designation, record.joining_date, record.leaving_date])

    sheet.append([])
    sheet.append(["Language Proficiencies"])
    sheet.append(["Language", "Speak", "Read", "Write"])
    for language in language_proficiencies:
        sheet.append([language.language, language.get_speak_display(), language.get_read_display(), language.get_write_display()])

    sheet.append([])
    sheet.append(["References"])
    sheet.append(["Name", "Occupation", "Phone Number", "Email"])
    for reference in references:
        sheet.append([reference.name, reference.occupation, reference.phone_number, reference.email])

    return workbook



################# Profile details views

# def profile_detail_view(request, profile_id):
#     # Get the specific profile by its ID
#     profile = get_object_or_404(Profile, id=profile_id)

#     all_profiles = Profile.objects.all() 

#     # Access related FamilyMember, EducationalQualification, etc.
#     family_members = profile.family_members.all()
#     educational_qualifications = profile.educational_qualifications.all()
#     employment_records = profile.employment_records.all()
#     language_proficiencies = profile.languages.all()
#     references = profile.references.all()


#     compliance_info = Employee.objects.filter(profile=profile).first()
#     # compliance_info = Employee.objects.filter(profile=profile).all()
#     print(compliance_info)

#     # Export to Excel if requested
#     if request.GET.get('export') == 'excel':
#         workbook = export_profile_to_excel(
#             profile,
#             family_members,
#             educational_qualifications,
#             employment_records,
#             language_proficiencies,
#             references
#         )

#         # Prepare the response
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename=Profile_{profile.Employee_Code}.xlsx'
#         workbook.save(response)
#         return response

#     # Pass the data to the template if not exporting
#     context = {
#         'profile': profile,
#         'all_profiles': all_profiles,
#         'family_members': family_members,
#         'educational_qualifications': educational_qualifications,
#         'employment_records': employment_records,
#         'language_proficiencies': language_proficiencies,
#         'references': references,
#         'compliance_info': compliance_info,
#     }

#     return render(request, 'profile_detail.html', context)



################################ update views for compliance data


def profile_detail_view(request, profile_id):
    # Get the specific profile by its ID
    profile = get_object_or_404(Profile, id=profile_id)
    all_profiles = Profile.objects.all()

    # Access related FamilyMember, EducationalQualification, etc.
    family_members = profile.family_members.all()
    educational_qualifications = profile.educational_qualifications.all()
    employment_records = profile.employment_records.all()
    language_proficiencies = profile.languages.all()
    references = profile.references.all()

    # Fetch the compliance data through the Applicant linked to this Profile
    applicant = getattr(profile.user, 'applicant', None)  # Access Applicant via the CustomUser relationship
    compliance_info = Employee.objects.filter(applicant=applicant).first() if applicant else None
    compensation_info = EmployeeCTC.objects.filter(applicant=applicant).first()
    # compensation_info = EmployeeCTC.objects.filter(profile=profile).first()

    # Export to Excel if requested
    if request.GET.get('export') == 'excel':
        workbook = export_profile_to_excel(
            # profile,
            # family_members,
            # educational_qualifications,
            # employment_records,
            # language_proficiencies,
            # references

            profile,
            family_members,
            educational_qualifications,
            employment_records,
            language_proficiencies,
            references,
            compliance_info,
            compensation_info  # Add these two arguments


        )

        # # Prepare the response
        # response = HttpResponse(
        #     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # )
        # response['Content-Disposition'] = f'attachment; filename=Profile_{profile.Employee_Code}.xlsx'
        # workbook.save(response)
        # return response


    # Prepare the response
        response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Profile_{profile.Employee_Code}.xlsx'
        workbook.save(response)
        return response


    # Pass the data to the template if not exporting
    context = {
        'profile': profile,
        'all_profiles': all_profiles,
        'family_members': family_members,
        'educational_qualifications': educational_qualifications,
        'employment_records': employment_records,
        'language_proficiencies': language_proficiencies,
        'references': references,
        'compliance_info': compliance_info,
        'compensation_info': compensation_info,
    }

    return render(request, 'profile_detail.html', context)




