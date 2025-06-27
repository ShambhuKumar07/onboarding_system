from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib.auth import logout
from .forms import CustomAuthenticationForm 
from django.contrib.auth import login
from compliance.models import Employee
from compensation.models import EmployeeCTC
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from users.models import Profile
# from employee.models import EmployeeCTC, Employee  # Assuming EmployeeCTC and Employee are in employee.models

def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            login(request, form.user_cache)
            try:
                if request.user.applicant:  # Trying to access the related applicant
                    return redirect('applicant:dashboard')
            except request.user._meta.model.applicant.RelatedObjectDoesNotExist:
                return redirect('core:main_page')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'users/login.html', {'form': form})

def logout_view(request):
    logout(request)

    return redirect('users:login_view')


# def export_profile_to_excel(profile, family_members, educational_qualifications, employment_records, language_proficiencies, references, compliance_info, compensation_info):
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Profile Details"

#     # Define headers for the personal details section
#     sheet.append(["Field", "Details"])
#     sheet.append(["Employee Code", profile.Employee_Code])
#     sheet.append(["First Name", profile.first_name])
#     sheet.append(["Middle Name", profile.middle_name])
#     sheet.append(["Last Name", profile.last_name])
#     sheet.append(["Designation", profile.designation])
#     sheet.append(["Date of Joining", profile.date_of_joining])
#     sheet.append(["End Date", profile.end_date]) 
#     sheet.append(["Email", profile.email])
#     sheet.append(["Phone Number", profile.phone_number])
#     sheet.append(["Gender", profile.get_gender_display()])
#     sheet.append(["Date of Birth", profile.date_of_birth])
#     sheet.append(["Present Address", profile.present_address])
#     sheet.append(["Permanent Address", profile.permanent_address])
#     sheet.append(["Personal Email", profile.personal_email])
#     sheet.append(["PAN Number", profile.pan_number])
#     sheet.append(["Marital Status", profile.marital_status])

#     # Adding a blank row
#     sheet.append([])

#     # Compliance section
#     sheet.append(["Compliance Information"])
#     if compliance_info:
#         sheet.append(["Compliance Field", "Compliance Value"])  # Add relevant fields here
 
#         sheet.append(["Compliance Status", getattr(compliance_info, 'status', 'N/A')])
#         sheet.append(["UAN Number", getattr(compliance_info, 'uan_number', 'N/A')])
#         sheet.append(["PAN Number", getattr(compliance_info, 'pan_number', 'N/A')])
#         sheet.append(["ESIC Number", getattr(compliance_info, 'esic_number', 'N/A')])
#     else:
#         sheet.append(["No compliance data available"])

#     sheet.append([])

#     # Compensation section
#     sheet.append(["Compensation Information"])
 

#     if compensation_info:
#         sheet.append(["Component", "Amount"])
#         salary = compensation_info.salary
#         contribution = compensation_info.contribution
#         deduction = compensation_info.deduction
#         sheet.append(["Basic Salary", salary.basic])
#         sheet.append(["HRA", salary.hra])
#         sheet.append(["Special Allowance", salary.special_allowance])
#         sheet.append(["Gross Salary", salary.gross_salary()])
#         sheet.append(["Total Contributions", contribution.emp_contribution()])
#         sheet.append(["Total Deductions", deduction.emp_deduction()])
#         sheet.append(["Net Salary (CTC)", compensation_info.calculate_ctc()])

#     else:
#         sheet.append(["No compensation data available"])

#     # Adding additional sections such as family members, educational qualifications, etc.
#     sheet.append([])
#     sheet.append(["Family Members"])
#     sheet.append(["Relation", "Name", "Date of Birth", "Sex", "Age"])
#     for member in family_members:
#         sheet.append([member.relation, member.name, member.date_of_birth, member.sex, member.age])

#     sheet.append([])
#     sheet.append(["Educational Qualifications"])
#     sheet.append(["Examination Passed", "Year of Passing", "School/College", "Division"])
#     for qualification in educational_qualifications:
#         sheet.append([qualification.examination_passed, qualification.year_of_passing, qualification.school_or_college, qualification.division])

#     sheet.append([])
#     sheet.append(["Employment Records"])
#     sheet.append(["Organization", "Designation", "Joining Date", "Leaving Date"])
#     for record in employment_records:
#         sheet.append([record.organization, record.designation, record.joining_date, record.leaving_date])

#     sheet.append([])
#     sheet.append(["Language Proficiencies"])
#     sheet.append(["Language", "Speak", "Read", "Write"])
#     for language in language_proficiencies:
#         sheet.append([language.language, language.get_speak_display(), language.get_read_display(), language.get_write_display()])

#     sheet.append([])
#     sheet.append(["References"])
#     sheet.append(["Name", "Occupation", "Phone Number", "Email"])
#     for reference in references:
#         sheet.append([reference.name, reference.occupation, reference.phone_number, reference.email])

#     return workbook



########################### Modifying code Personal Information for columns wise


def export_profile_to_excel(profile, family_members, educational_qualifications, employment_records, language_proficiencies, references, compliance_info, compensation_info):
    import openpyxl
    from openpyxl.styles import Alignment

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Profile Details"

    # Define headers for the personal details section
    personal_headers = [
        "Employee Code", "First Name", "Middle Name", "Last Name",
        "Designation", "Date of Joining", "End Date", "Email",
        "Phone Number", "Gender", "Date of Birth", "Present Address",
        "Permanent Address", "Personal Email", "PAN Number", "Marital Status"
    ]
    personal_details = [
        profile.Employee_Code, profile.first_name, profile.middle_name,
        profile.last_name, profile.designation, profile.date_of_joining,
        profile.end_date, profile.email, profile.phone_number,
        profile.get_gender_display(), profile.date_of_birth,
        profile.present_address, profile.permanent_address,
        profile.personal_email, profile.pan_number, profile.marital_status
    ]

    # Write headers and personal details in column-wise format
    for col, header in enumerate(personal_headers, start=1):
        sheet.cell(row=1, column=col, value=header)  # Header row
        sheet.cell(row=2, column=col, value=personal_details[col - 1])  # Data row

    # Adjust alignment for readability
    for col in range(1, len(personal_headers) + 1):
        sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        sheet.cell(row=2, column=col).alignment = Alignment(horizontal='center')

    # Add additional sections for family members, education, etc.
    row_offset = 4

    # Family Members Section
    sheet.cell(row=row_offset, column=1, value="Family Members")
    row_offset += 1
    sheet.append(["Relation", "Name", "Date of Birth", "Sex", "Age"])
    for member in family_members:
        sheet.append([member.relation, member.name, member.date_of_birth, member.sex, member.age])
    row_offset += len(family_members) + 2

    # Educational Qualifications
    sheet.cell(row=row_offset, column=1, value="Educational Qualifications")
    row_offset += 1
    sheet.append(["Examination Passed", "Year of Passing", "School/College", "Division"])
    for qualification in educational_qualifications:
        sheet.append([
            qualification.examination_passed,
            qualification.year_of_passing,
            qualification.school_or_college,
            qualification.division,
        ])
    row_offset += len(educational_qualifications) + 2

    # Employment Records
    sheet.cell(row=row_offset, column=1, value="Employment Records")
    row_offset += 1
    sheet.append(["Organization", "Designation", "Joining Date", "Leaving Date"])
    for record in employment_records:
        sheet.append([
            record.organization, record.designation,
            record.joining_date, record.leaving_date
        ])
    row_offset += len(employment_records) + 2

    # Language Proficiencies
    sheet.cell(row=row_offset, column=1, value="Languages Known")
    row_offset += 1
    sheet.append(["Language", "Speak", "Read", "Write"])
    for language in language_proficiencies:
        sheet.append([
            language.language,
            language.get_speak_display(),
            language.get_read_display(),
            language.get_write_display()
        ])
    row_offset += len(language_proficiencies) + 2

    # References
    sheet.cell(row=row_offset, column=1, value="References")
    row_offset += 1
    sheet.append(["Name", "Occupation", "Phone Number", "Email"])
    for reference in references:
        sheet.append([
            reference.name, reference.occupation,
            reference.phone_number, reference.email
        ])
    row_offset += len(references) + 2

    # Compliance Information
    sheet.cell(row=row_offset, column=1, value="Compliance Information")
    row_offset += 1
    if compliance_info:
        sheet.append(["UAN Number", compliance_info.uan_number])
        sheet.append(["PAN Number", compliance_info.pan_number])
        sheet.append(["ESIC Number", compliance_info.esic_number])
    else:
        sheet.append(["No compliance data available"])
    row_offset += 3

    # Compensation Information
    sheet.cell(row=row_offset, column=1, value="Compensation Information")
    row_offset += 1
    if compensation_info:
        sheet.append(["Component", "Amount"])
        salary = compensation_info.salary
        contribution = compensation_info.contribution
        deduction = compensation_info.deduction
        sheet.append(["Basic Salary", salary.basic])
        sheet.append(["HRA", salary.hra])
        sheet.append(["Special Allowance", salary.special_allowance])
        sheet.append(["Gross Salary", salary.gross_salary()])
        sheet.append(["Total Contributions", contribution.emp_contribution()])
        sheet.append(["Total Deductions", deduction.emp_deduction()])
        sheet.append(["Net Salary (CTC)", compensation_info.calculate_ctc()])
    else:
        sheet.append(["No compensation data available"])

    return workbook





################################ update views for compliance data

def profile_detail_view(request, profile_id):
    # Get the specific profile by its ID
    profile = get_object_or_404(Profile, id=profile_id)
    all_profiles = Profile.objects.all()

    # Access related FamilyMember, EducationalQualification, etc.
    family_members = profile.family_members.all()
    educational_qualifications = profile.educational_qualifications.all()
    employment_records = profile.employment_records.all()
    language_proficiencies = profile.languages.all()
    references = profile.references.all()

    # Fetch the compliance data through the Applicant linked to this Profile
    applicant = getattr(profile.user, 'applicant', None)  # Access Applicant via the CustomUser relationship
    compliance_info = Employee.objects.filter(applicant=applicant).first() if applicant else None
    compensation_info = EmployeeCTC.objects.filter(applicant=applicant).first()
    # compensation_info = EmployeeCTC.objects.filter(profile=profile).first()

    # Export to Excel if requested
    if request.GET.get('export') == 'excel':
        workbook = export_profile_to_excel(


            profile,
            family_members,
            educational_qualifications,
            employment_records,
            language_proficiencies,
            references,
            compliance_info,
            compensation_info  # Add these two arguments
        )

    # Prepare the response
        response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Profile_{profile.Employee_Code}.xlsx'
        workbook.save(response)
        return response


    # Pass the data to the template if not exporting
    context = {
        'profile': profile,
        'all_profiles': all_profiles,
        'family_members': family_members,
        'educational_qualifications': educational_qualifications,
        'employment_records': employment_records,
        'language_proficiencies': language_proficiencies,
        'references': references,
        'compliance_info': compliance_info,
        'compensation_info': compensation_info,
    }

    return render(request, 'profile_detail.html', context)



# from django.shortcuts import render
# from users.models import Profile

# def attrition_report_view(request):
#     profiles = Profile.objects.filter(end_date__isnull=False)  # Only include profiles with an end_date
#     return render(request, 'users/attrition_report.html', {'profiles': profiles})



from django.shortcuts import render
from applicant.models import Applicant
from users.models import Profile
from django.utils.dateparse import parse_date

# def hiring_report(request):
#     # joining_date = request.GET.get('joining_date')
#     from_date = request.GET.get('date_of_joining')
#     end_date = request.GET.get('date_of_joining')
#     profiles = Applicant.objects.all()
#     # profiles = Profile.objects.filter(end_date__isnull=False)  # Only include profiles with an end_date
    
#     # if joining_date:
#     #     profiles = profiles.filter(date_of_joining__gte=parse_date(joining_date))
#     # # if end_date:
#     # #     profiles = profiles.filter(end_date__lte=parse_date(end_date))
#     # Apply the filter based only on the 'end_date' range
#     if from_date and end_date:
#           profiles = profiles.filter(
#         Hiring_date__range=(parse_date(from_date), parse_date(end_date))
#         )
#     elif from_date:
#         profiles = profiles.filter(end_date__gte=parse_date(from_date))
#     elif end_date:
#         profiles = profiles.filter(end_date__lte=parse_date(end_date))      

#     return render(request, 'users/hiring_report.html', {'profiles': profiles, 'from_date': from_date, 'end_date': end_date})

def hiring_report(request):
    from_date = request.GET.get('from_date')
    date_of_joining = request.GET.get('date_of_joining')
    
    profiles = Profile.objects.all()
    
    # Apply the filter based on 'date_of_joining' range
    if from_date and date_of_joining:
        profiles = profiles.filter(
            date_of_joining__range=(parse_date(from_date), parse_date(date_of_joining))
        )
    elif from_date:
        profiles = profiles.filter(date_of_joining__gte=parse_date(from_date))
    elif date_of_joining:
        profiles = profiles.filter(date_of_joining__lte=parse_date(date_of_joining))
    
    return render(request, 'users/hiring_report.html', {'profiles': profiles, 'from_date': from_date, 'date_of_joining': date_of_joining})

from django.shortcuts import render
from users.models import Profile
from django.utils.dateparse import parse_date
from django.db.models import Q

def attrition_report(request):
    from_date = request.GET.get('from_date')
    end_date = request.GET.get('end_date')
    
    profiles = Profile.objects.filter(end_date__isnull=False)  # Only include profiles with an end_date

    # Apply the filter based only on the 'end_date' range
    if from_date and end_date:
        profiles = profiles.filter(
            end_date__range=(parse_date(from_date), parse_date(end_date))
        )
    elif from_date:
        profiles = profiles.filter(end_date__gte=parse_date(from_date))
    elif end_date:
        profiles = profiles.filter(end_date__lte=parse_date(end_date))
    
    return render(request, 'users/attrition_report.html', {'profiles': profiles, 'from_date': from_date, 'end_date': end_date})


#################### modifying the code Column wise#################

 


# def attrition_report(request):
#     import datetime
#     import csv
#     from django.http import HttpResponse
#     from users.models import Profile
#     from django.utils.dateparse import parse_date

#     from_date = request.GET.get('from_date')
#     end_date = request.GET.get('end_date')

#     # Filter profiles with an end_date
#     profiles = Profile.objects.filter(end_date__isnull=False)

#     # Apply the filter based on the 'end_date' range
#     if from_date and end_date:
#         profiles = profiles.filter(
#             end_date__range=(parse_date(from_date), parse_date(end_date))
#         )
#     elif from_date:
#         profiles = profiles.filter(end_date__gte=parse_date(from_date))
#     elif end_date:
#         profiles = profiles.filter(end_date__lte=parse_date(end_date))

#     # Check if the user wants to export the data as Excel
#     if request.GET.get('export') == 'excel':
#         # Create a response object for the CSV file
#         response = HttpResponse(content_type='text/csv')
#         filename = f"attrition_report_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'

#         # Write data to the CSV file
#         writer = csv.writer(response)

#         # Write column headers
#         headers = [
#             'Field', 'Employee Code', 'First Name', 'Middle Name', 'Last Name',
#             'Designation', 'Date of Joining', 'End Date', 'Email',
#             'Phone Number', 'Gender', 'Date of Birth', 'Present Address',
#             'Permanent Address', 'Personal Email', 'PAN Number', 'Marital Status'
#         ]
#         writer.writerow(headers)

#         # Write the profile details (row-wise layout)
#         for profile in profiles:
#             row = [
#                 'Details', profile.Employee_Code, profile.first_name,
#                 profile.middle_name or '', profile.last_name, profile.designation,
#                 profile.date_of_joining, profile.end_date, profile.email,
#                 profile.phone_number, profile.get_gender_display(),
#                 profile.date_of_birth, profile.present_address,
#                 profile.permanent_address, profile.personal_email,
#                 profile.pan_number, profile.marital_status
#             ]
#             writer.writerow(row)

#         return response

#     # Render the HTML template
#     return render(request, 'users/attrition_report.html', {'profiles': profiles, 'from_date': from_date, 'end_date': end_date})





########################### export all applicant details


'''def export_all_profiles_to_excel(profiles):
    pass
    import openpyxl
    from openpyxl.styles import Alignment

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "All Profiles Details"

    # Define headers
    headers = [
        # Personal Information
        "Employee Code", "First Name", "Middle Name", "Last Name", "Designation",
        "Date of Joining", "End Date", "Email", "Phone Number", "Gender",
        "Date of Birth", "Present Address", "Permanent Address", "Personal Email",
        "PAN Number", "Marital Status",
        # Bank Information
        "Bank Name", "IFSC Code", "Bank Account Number",
        # Compensation Details
        "Basic Salary", "HRA", "Special Allowance", "Gross Salary", 
        "Total Contributions", "Total Deductions", "Net Salary (CTC)",
        # Compliance Information
        "UAN Number", "ESIC Number",
        # Emergency Contact Information
        "Emergency Contact Name", "Relation", "Contact Number",
        # Family Members
        "Family Members (Relation | Name | DOB | Gender | Age)",
        # Educational Qualifications
        "Educational Qualifications (Exam Passed | Year | School | Division)",
        # Employment Records
        "Employment Records (Organization | Designation | Joining | Leaving)",
        # Languages Known
        "Languages Known (Language | Speak | Read | Write)",
        # References
        "References (Name | Occupation | Phone | Email)",
    ]

    # Write headers to the first row
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
        sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')

    # Populate rows with profile data
    for row_idx, profile in enumerate(profiles, start=2):
        data = [
            # Personal Information
            profile.Employee_Code, profile.first_name, profile.middle_name,
            profile.last_name, profile.designation, profile.date_of_joining,
            profile.end_date, profile.email, profile.phone_number,
            profile.get_gender_display(), profile.date_of_birth,
            profile.present_address, profile.permanent_address,
            profile.personal_email, profile.pan_number, profile.marital_status,
            # Bank Information
            profile.bank_name, profile.ifsc_code, profile.bank_account_number,
        ]

        # Add compensation details if available
        if hasattr(profile, 'compensation_info') and profile.compensation_info:
            salary = profile.compensation_info.salary
            contribution = profile.compensation_info.contribution
            deduction = profile.compensation_info.deduction
            data.extend([
                salary.basic, salary.hra, salary.special_allowance,
                salary.gross_salary(), contribution.emp_contribution(),
                deduction.emp_deduction(), profile.compensation_info.calculate_ctc()
            ])
        else:
            data.extend(["N/A"] * 7)

        # Add compliance information
        compliance_info = getattr(profile, 'compliance_info', None)
        if compliance_info:
            data.extend([compliance_info.uan_number, compliance_info.esic_number])
        else:
            data.extend(["N/A", "N/A"])

        # Add emergency contact information
        data.extend([
            profile.emergency_contact_name,
            profile.emergency_contact_relation,
            profile.emergency_contact_number,
        ])

        # Add family members
        family_members = profile.family_members.all()
        family_data = " | ".join(
            f"{member.relation} | {member.name} | {member.date_of_birth} | {member.sex} | {member.age}"
            for member in family_members
        )
        data.append(family_data or "N/A")

        # Add educational qualifications
        qualifications = profile.educational_qualifications.all()
        education_data = " | ".join(
            f"{q.examination_passed} | {q.year_of_passing} | {q.school_or_college} | {q.division}"
            for q in qualifications
        )
        data.append(education_data or "N/A")

        # Add employment records
        employment_records = profile.employment_records.all()
        employment_data = " | ".join(
            f"{e.organization} | {e.designation} | {e.joining_date} | {e.leaving_date}"
            for e in employment_records
        )
        data.append(employment_data or "N/A")

        # Add languages known
        languages = profile.language_proficiencies.all()
        language_data = " | ".join(
            f"{lang.language} | {lang.get_speak_display()} | {lang.get_read_display()} | {lang.get_write_display()}"
            for lang in languages
        )
        data.append(language_data or "N/A")

        # Add references
        references = profile.references.all()
        reference_data = " | ".join(
            f"{ref.name} | {ref.occupation} | {ref.phone_number} | {ref.email}"
            for ref in references
        )
        data.append(reference_data or "N/A")

        # Write data to the sheet
        for col_idx, value in enumerate(data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    return workbook
def export_all_profiles_to_excel(profiles):
    import openpyxl
    from openpyxl.styles import Alignment

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "All Profiles Details"

    # Define headers
    headers = [
        # Personal Information
        "Employee Code", "First Name", "Middle Name", "Last Name", "Designation",
        "Date of Joining", "End Date", "Email", "Phone Number", "Gender",
        "Date of Birth", "Present Address", "Permanent Address", "Personal Email",
        "PAN Number", "Marital Status",
        # Bank Information
        "Bank Name", "IFSC Code", "Bank Account Number",
        # Compensation Details
        "Basic Salary", "HRA", "Special Allowance", "Gross Salary", 
        "Total Contributions", "Total Deductions", "Net Salary (CTC)",
        # Compliance Information
        "UAN Number", "ESIC Number",
        # Emergency Contact Information
        "Emergency Contact Name", "Relation", "Contact Number",
        # Family Members
        "Family Members (Relation | Name | DOB | Gender | Age)",
        # Educational Qualifications
        "Educational Qualifications (Exam Passed | Year | School | Division)",
        # Employment Records
        "Employment Records (Organization | Designation | Joining | Leaving)",
        # Languages Known
        "Languages Known (Language | Speak | Read | Write)",
        # References
        "References (Name | Occupation | Phone | Email)",
    ]

    # Write headers to the first row
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
        sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')

    # Populate rows with profile data
    for row_idx, profile in enumerate(profiles, start=2):
        data = [
            # Personal Information
            profile.Employee_Code, profile.first_name, profile.middle_name,
            profile.last_name, profile.designation, profile.date_of_joining,
            profile.end_date, profile.email, profile.phone_number,
            profile.get_gender_display(), profile.date_of_birth,
            profile.present_address, profile.permanent_address,
            profile.personal_email, profile.pan_number, profile.marital_status,
            # Bank Information
            profile.bank_name, profile.ifsc_code, profile.bank_account_number,
        ]

        # Add compensation details if available
        if hasattr(profile, 'compensation_info') and profile.compensation_info:
            salary = profile.compensation_info.salary
            contribution = profile.compensation_info.contribution
            deduction = profile.compensation_info.deduction
            data.extend([
                salary.basic, salary.hra, salary.special_allowance,
                salary.gross_salary(), contribution.emp_contribution(),
                deduction.emp_deduction(), profile.compensation_info.calculate_ctc()
            ])
        else:
            data.extend(["N/A"] * 7)

        # Add compliance information
        compliance_info = getattr(profile, 'compliance_info', None)
        if compliance_info:
            data.extend([compliance_info.uan_number, compliance_info.esic_number])
        else:
            data.extend(["N/A", "N/A"])

        # Add emergency contact information
        data.extend([
            profile.emergency_contact_name,
            profile.emergency_contact_relation,
            profile.emergency_contact_number,
        ])

        # Add family members
        family_members = profile.family_members.all()
        family_data = " | ".join(
            f"{member.relation} | {member.name} | {member.date_of_birth} | {member.sex} | {member.age}"
            for member in family_members
        )
        data.append(family_data or "N/A")

        # Add educational qualifications
        qualifications = profile.educational_qualifications.all()
        education_data = " | ".join(
            f"{q.examination_passed} | {q.year_of_passing} | {q.school_or_college} | {q.division}"
            for q in qualifications
        )
        data.append(education_data or "N/A")

        # Add employment records
        employment_records = profile.employment_records.all()
        employment_data = " | ".join(
            f"{e.organization} | {e.designation} | {e.joining_date} | {e.leaving_date}"
            for e in employment_records
        )
        data.append(employment_data or "N/A")

        # Add languages known
        languages = profile.language_proficiencies.all()
        language_data = " | ".join(
            f"{lang.language} | {lang.get_speak_display()} | {lang.get_read_display()} | {lang.get_write_display()}"
            for lang in languages
        )
        data.append(language_data or "N/A")

        # Add references
        references = profile.references.all()
        reference_data = " | ".join(
            f"{ref.name} | {ref.occupation} | {ref.phone_number} | {ref.email}"
            for ref in references
        )
        data.append(reference_data or "N/A")

        # Write data to the sheet
        for col_idx, value in enumerate(data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    return workbook'''

    
# def export_all_applicants_to_mis(profiles):
#     import openpyxl
#     from openpyxl.styles import Alignment, Font

#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicant MIS Report"

#     # Define headers
#     headers = [
#         # Personal Information
#         "Applicant ID", "First Name", "Middle Name", "Last Name", "Email", "Phone Number",
#         "Gender", "Date of Birth", "Current Address", "Permanent Address",
#         # Application Details
#         "Position Applied For", "Application Date", "Application Status", 
#         "Interview Date", "Interview Status", "Remarks",
#         # Educational Qualifications
#         "Educational Qualifications (Degree | Year | Institution | Grade)",
#         # Work Experience
#         "Work Experience (Company | Designation | From | To | Responsibilities)",
#         # Skills
#         "Skills (Skill | Proficiency Level)",
#         # Certifications
#         "Certifications (Certification | Institution | Year Obtained)",
#         # References
#         "References (Name | Occupation | Contact Details)",
#         # Emergency Contact Information
#         "Emergency Contact Name", "Relation", "Contact Number"
#     ]

#     # Write headers to the first row
#     for col, header in enumerate(headers, start=1):
#         cell = sheet.cell(row=1, column=col, value=header)
#         cell.alignment = Alignment(horizontal='center', vertical='center')
#         cell.font = Font(bold=True)

#     # Populate rows with applicant data
#     for row_idx, profile in enumerate(profiles, start=2):
#         data = [
#             # Personal Information
#             getattr(profile, 'applicant_id', 'N/A'),
#             getattr(profile, 'first_name', 'N/A'),
#             getattr(profile, 'middle_name', 'N/A'),
#             getattr(profile, 'last_name', 'N/A'),
#             getattr(profile, 'email', 'N/A'),
#             getattr(profile, 'phone_number', 'N/A'),
#             getattr(profile, 'get_gender_display', lambda: 'N/A')(),
#             getattr(profile, 'date_of_birth', 'N/A'),
#             getattr(profile, 'current_address', 'N/A'),
#             getattr(profile, 'permanent_address', 'N/A'),
#             # Application Details
#             getattr(profile, 'position_applied_for', 'N/A'),
#             getattr(profile, 'application_date', 'N/A'),
#             getattr(profile, 'application_status', 'N/A'),
#             getattr(profile, 'interview_date', 'N/A'),
#             getattr(profile, 'interview_status', 'N/A'),
#             getattr(profile, 'remarks', 'N/A'),
#         ]

#         # Educational Qualifications
#         qualifications = getattr(profile, 'educational_qualifications', None)
#         education_data = " | ".join(
#             f"{q.degree} | {q.year_of_passing} | {q.institution} | {q.grade}"
#             for q in qualifications.all()
#         ) if qualifications else "N/A"
#         data.append(education_data)

#         # Work Experience
#         work_experiences = getattr(profile, 'work_experiences', None)
#         experience_data = " | ".join(
#             f"{exp.company} | {exp.designation} | {exp.start_date} | {exp.end_date} | {exp.responsibilities}"
#             for exp in work_experiences.all()
#         ) if work_experiences else "N/A"
#         data.append(experience_data)

#         # Skills
#         skills = getattr(profile, 'skills', None)
#         skill_data = " | ".join(
#             f"{skill.name} | {skill.proficiency_level}"
#             for skill in skills.all()
#         ) if skills else "N/A"
#         data.append(skill_data)

#         # Certifications
#         certifications = getattr(profile, 'certifications', None)
#         certification_data = " | ".join(
#             f"{cert.name} | {cert.institution} | {cert.year_obtained}"
#             for cert in certifications.all()
#         ) if certifications else "N/A"
#         data.append(certification_data)

#         # References
#         references = getattr(profile, 'references', None)
#         reference_data = " | ".join(
#             f"{ref.name} | {ref.occupation} | {ref.contact_details}"
#             for ref in references.all()
#         ) if references else "N/A"
#         data.append(reference_data)

#         # Emergency Contact Information
#         data.extend([
#             getattr(profile, 'emergency_contact_name', 'N/A'),
#             getattr(profile, 'emergency_contact_relation', 'N/A'),
#             getattr(profile, 'emergency_contact_number', 'N/A'),
#         ])

#         # Write data to the sheet
#         for col_idx, value in enumerate(data, start=1):
#             sheet.cell(row=row_idx, column=col_idx, value=value)

#     return workbook

# from django.http import HttpResponse
# from io import BytesIO

# def export_mis_report(request):
#     profiles = Applicant.objects.all()  # Adjust this query to fetch the relevant data
#     workbook = export_all_applicants_to_mis(profiles)

#     # Save workbook to an in-memory buffer
#     buffer = BytesIO()
#     workbook.save(buffer)
#     buffer.seek(0)

#     # Create an HTTP response with the buffer content
#     response = HttpResponse(
#         buffer,
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="Applicant_MIS_Report.xlsx"'
#     return response
'''from django.http import HttpResponse
from io import BytesIO
from openpyxl import Workbook  # Explicit import for clarity

def export_all_applicants_to_mis(profiles):
    import openpyxl
    from openpyxl.styles import Alignment, Font

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Applicant MIS Report"

    # Define headers
    headers = [
        "Applicant ID", "First Name", "Middle Name", "Last Name", "Email", "Phone Number",
        "Gender", "Date of Birth", "Current Address", "Permanent Address",
        "Position Applied For", "Application Date", "Application Status", 
        "Interview Date", "Interview Status", "Remarks",
        "Educational Qualifications (Degree | Year | Institution | Grade)",
        "Work Experience (Company | Designation | From | To | Responsibilities)",
        "Skills (Skill | Proficiency Level)",
        "Certifications (Certification | Institution | Year Obtained)",
        "References (Name | Occupation | Contact Details)",
        "Emergency Contact Name", "Relation", "Contact Number"
    ]

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)

    # Populate rows
    for row_idx, profile in enumerate(profiles, start=2):
        data = [
            getattr(profile, 'applicant_id', 'N/A'),
            getattr(profile, 'first_name', 'N/A'),
            getattr(profile, 'middle_name', 'N/A'),
            getattr(profile, 'last_name', 'N/A'),
            getattr(profile, 'email', 'N/A'),
            getattr(profile, 'phone_number', 'N/A'),
            getattr(profile, 'get_gender_display', lambda: 'N/A')(),
            getattr(profile, 'date_of_birth', 'N/A'),
            getattr(profile, 'current_address', 'N/A'),
            getattr(profile, 'permanent_address', 'N/A'),
            getattr(profile, 'position_applied_for', 'N/A'),
            getattr(profile, 'application_date', 'N/A'),
            getattr(profile, 'application_status', 'N/A'),
            getattr(profile, 'interview_date', 'N/A'),
            getattr(profile, 'interview_status', 'N/A'),
            getattr(profile, 'remarks', 'N/A'),
        ]

        # Add the rest of the data dynamically as in your original code

        for col_idx, value in enumerate(data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    return workbook


def export_mis_report(request):
    profiles = Applicant.objects.all()  # Adjust this query to fetch the relevant data
    workbook = export_all_applicants_to_mis(profiles)

    # Save workbook to an in-memory buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Create an HTTP response with the buffer content
    response = HttpResponse(
        buffer.getvalue(),  # Use .getvalue() to retrieve bytes from the buffer
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Applicant_MIS_Report.xlsx"'
    buffer.close()  # Close the buffer to free resources
    return response'''

from django.http import HttpResponse
from io import BytesIO
import openpyxl

def export_all_applicants_to_mis(profiles):
    from openpyxl.styles import Alignment, Font

    # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Applicant MIS Report"

    # Define headers
    headers = [
        "Applicant ID", "First Name", "Middle Name", "Last Name", "Email", "Phone Number",
        "Gender", "Date of Birth", "Current Address", "Permanent Address",
        "Position Applied For", "Application Date", "Application Status", 
        "Interview Date", "Interview Status", "Remarks",
    ]

    # Write headers to the first row
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

    # Populate rows with applicant data
    for row_idx, profile in enumerate(profiles, start=2):
        data = [
            getattr(profile, 'applicant_id', 'N/A'),
            getattr(profile, 'first_name', 'N/A'),
            getattr(profile, 'middle_name', 'N/A'),
            getattr(profile, 'last_name', 'N/A'),
            getattr(profile, 'email', 'N/A'),
            getattr(profile, 'phone_number', 'N/A'),
            getattr(profile, 'get_gender_display', lambda: 'N/A')(),
            getattr(profile, 'date_of_birth', 'N/A'),
            getattr(profile, 'current_address', 'N/A'),
            getattr(profile, 'permanent_address', 'N/A'),
            getattr(profile, 'position_applied_for', 'N/A'),
            getattr(profile, 'application_date', 'N/A'),
            getattr(profile, 'application_status', 'N/A'),
            getattr(profile, 'interview_date', 'N/A'),
            getattr(profile, 'interview_status', 'N/A'),
            getattr(profile, 'remarks', 'N/A'),
        ]

        # Write data to the sheet
        for col_idx, value in enumerate(data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    return workbook

def export_mis_report(request):
    # Fetch applicant data (adjust query as needed)
    profiles = Applicant.objects.all()

    # Generate the workbook
    workbook = export_all_applicants_to_mis(profiles)

    # Save the workbook to an in-memory buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Create and return the HTTP response
    response = HttpResponse(
        buffer.getvalue(),  # Get the binary content of the buffer
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="Applicant_MIS_Report.xlsx"'
    buffer.close()  # Free resources by closing the buffer
    return response

